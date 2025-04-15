import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

# Tkinterのウィンドウを非表示にする
root = tk.Tk()
root.withdraw()

# ファイルダイアログを表示してファイルパスを取得
file_path = filedialog.askopenfilename(
    title="Excelファイルを選択してください",
    filetypes=[("Excel files", "*.xlsx *.xlsm")]
)

# ファイルが選択された場合のみ処理
if file_path:
    wb = load_workbook(file_path)
    ws = wb.active
    print("選択されたファイル：", file_path)
    print("先頭セルの値：", ws.cell(row=1, column=1).value)
else:
    print("ファイルが選択されませんでした")
