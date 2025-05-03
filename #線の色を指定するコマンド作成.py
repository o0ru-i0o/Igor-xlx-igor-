#線の色を指定するコマンド作成
import openpyxl;
import os;    # osモジュールをインポート



wb = openpyxl.load_workbook(f'{os.path.dirname(__file__)}/個人用設定.xlsx');    # 個人用設定.xlsxを読み込む
ws = wb["TraceColor"];    # TraceColorを取得

min_row = 2;    # 開始行
max_row = ws.max_row;    # 最大行数を取得
min_col = 1;    # 開始列
max_col = 4;    # 最大列数を取得
Trace_Color_Settings= ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True);    # TraceColorの設定を取得

print(f'{Trace_Color_Settings=}');
#Trace_Color_Settings_tuple = tuple(Trace_Color_Settings);    # Trace_Color_Settingsをタプルに変換
#Trace_Color_Settings_dict  = list(Trace_Color_Settings);   #意味ないみたい
#print(f'{tuple=}');
#print(f'{Trace_Color_Settings_tuple[0]=}');
#print(f'{Trace_Color_Settings_tuple[0][0]=}');
#print(f'{Trace_Color_Settings_tuple[0][0].value=}');
#print(f'{Trace_Color_Settings[0]}');

#↓だとlist(Trace_Color_Settings_mz)=[(1,), (2,), (3,)]になっちゃう
#Trace_Color_Settings_mz = ws.iter_rows(min_row=2, max_row=max_row, max_col=1, values_only=True);
Trace_Color_Settings_mz = ws["A"];    # TraceColorの設定を取得
Trace_Color_Settings_Color = ws.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=4, values_only=True);

Trace_Color_Settings_mz_list = [cell.value for cell in Trace_Color_Settings_mz if type(cell.value) == int];


Trace_Color_Settings_Color_tuple = tuple(Trace_Color_Settings_Color);    # Trace_Color_Settingsをタプルに変換
print(f'{Trace_Color_Settings_Color_tuple=}');

#Trace_Color_Settings_dict
#print(f'{Trace_Color_Settings_dict=}');

#これだと上手くいかないなぁ
#Trace_Color_Settings_dict = dict(zip(Trace_Color_Settings_mz_list, Trace_Color_Settings_Color));    # Trace_Color_Settingsを辞書に変換

Trace_Color_Settings_dict = {mz:color for mz, color in zip(Trace_Color_Settings_mz_list, Trace_Color_Settings_Color_tuple) };    # Trace_Color_Settingsを辞書に変換

print((Trace_Color_Settings_dict));

print(f'{Trace_Color_Settings_dict[2]=}');
print(f'{Trace_Color_Settings_dict[1][0]=}');

"""
for i in range(len(Trace_Color_Settings_tuple)):
    if i == 2:
        print(f'Trace Color of m/z is {Trace_Color_Settings_tuple[i][1]}');
"""



