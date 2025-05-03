import openpyxl;
import os;
import os.path;
import tkinter;
import tkinter.filedialog;
import pyperclip
import csv

import re
import sys

from tkinter import messagebox
import traceback

#グローバル変数の定義
wb = None;
sheet_names = None;
ws = None;
file_path = None;
mass_number = None;
csv_file_path_with_collon = None;

import pandas
import chardet
import os
import tkinter.filedialog

def csv_to_excel_with_pandas(path=None):
    #global wb;
    #global sheet_names;
    #global ws;
    global file_path;
    
    file_path = path;    # グローバル変数にファイルパスを格納


    # ファイル選択
    csv_file_path = tkinter.filedialog.askopenfilename(
        title="CSVファイルを選んでね！",
        filetypes=[("CSV files", "*.csv;*.CSV")]
    )

    if not csv_file_path:
        notify_user("キャンセルされたよ〜");
        print("キャンセルされたよ〜")
        return

    # 文字コードを自動検出
    with open(csv_file_path, 'rb') as f:
        raw_data = f.read()
        result = chardet.detect(raw_data)
        detected_encoding = result['encoding']
        #notify_user(f"検出された文字コード：{detected_encoding}")
        print(f"✅ 検出された文字コード：{detected_encoding}")

    # pandasで読み込んで → Excelに出力！
    try:
        df = pandas.read_csv(csv_file_path, encoding=detected_encoding, engine='python', on_bad_lines='skip')

        # 拡張子を安全に置き換え
        filename_root, _ = os.path.splitext(csv_file_path)
        excel_file_path = filename_root + ".xlsx"

        df.to_excel(excel_file_path, index=False)
        #notify_user(f"✅ pandasで変換完了！: {excel_file_path}")
        print(f"✅ pandasで変換完了！: {excel_file_path}")
    except Exception as e:
        #notify_user(f"❌ pandasでの読み込みエラー：{e}")
        print("❌ pandasでの読み込みエラー：", e)

    file_path = excel_file_path;    # グローバル変数にファイルパスを格納

def csv_to_excel_by_csvreader():
    global wb;
    global sheet_names;
    global ws;
    global file_path;

    # CSVファイルの読み込み
    # ファイルダイアログを表示してCSVファイルパスを取得
    csv_file_path = tkinter.filedialog.askopenfilename(
        title="CSVファイルを選択してください",
        filetypes=[("CSV files", "*.csv")]
    )
    excel_file_path = csv_file_path.replace(".csv",".xlsx")
    file_path = excel_file_path;    # グローバル変数にファイルパスを格納


    if csv_file_path:
        csv_file_path = csv_file_path.replace("C:", "");
        csv_file_path = csv_file_path.replace("D:", "");


        wb = openpyxl.Workbook();
        print("選択されたファイル：", csv_file_path);
        ws = wb.active

        

        with open(csv_file_path) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)

        # 出力先のExcelファイル名を生成
        excel_file_path = csv_file_path.replace(".csv","_convert.xlsx")


        wb.save(excel_file_path)
    
    else:
        print("CSVファイルが選択されませんでした。")
def csv_to_excel_test():
    global wb;
    global sheet_names;
    global ws;
    global file_path;

    # ファイルダイアログを表示してCSVファイルパスを取得
    csv_file_path = tkinter.filedialog.askopenfilename(
        title="CSVファイルを選択してください",
        filetypes=[("CSV files", "*.csv")]
    )

    # ファイルが選択された場合のみ処理
    if csv_file_path:
        # 出力先のExcelファイル名を生成
        excel_file_path = os.path.splitext(csv_file_path)[0] + "_converted.xlsx"

        # CSVファイルの読み込みと置換
        with open(csv_file_path, 'r', newline='', encoding='utf-8') as file, \
                open('file_out.csv', 'w', newline='', encoding='utf-8') as fileout:
            text = re.sub(r'\s* ', ',', file.read())
            print(text, file=fileout)
            print('置換完了')

        # CSVファイルの読み込み
        data = pandas.read_csv('file_out.csv', encoding='utf-8')

        # Excel形式で出力
        data.to_excel(excel_file_path, encoding='utf-8', index=False)

        print(f'CSV > Excel変換完了: {excel_file_path}')

        file_path = excel_file_path;    # グローバル変数にファイルパスを格納
    else:
        print("CSVファイルが選択されませんでした。")

# Tkinterのウィンドウを非表示にする
#root = tkinter.Tk();
#root.withdraw();


def read_excel_file(path=None, progress_callback=None):
    global wb;
    global sheet_names;
    global ws;
    global file_path;

    file_path = path;    # グローバル変数にファイルパスを格納
    """
    # ファイルダイアログを表示してファイルパスを取得
    file_path = tkinter.filedialog.askopenfilename(
        title="Excelファイルを選択してください",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )
    """

    # ファイルが選択された場合のみ処理
    if file_path:
        # Excelファイルを読み込む
        wb = openpyxl.load_workbook(file_path);
        print("選択されたファイル：", file_path);

        sheet_names = wb.sheetnames;    # シート名のリストを取得
        
        notify_user(f"{str(file_path)}'\n を読み込みます")
        print(f"選択されたファイル：{file_path}");    # 選択されたファイル名を表示
            
        if progress_callback:
            progress_callback(10)


        for i, sheet_name in enumerate(sheet_names):
            ws = wb[sheet_name];    # シートを取得
            print(f"{i+1}番目のシート名：{sheet_name}");    # シート名を表示
            print("先頭セルの値：", ws.cell(row=1, column=1).value)
            print(f"最大行数：{ws.max_row}");    # A列の行数を表示
          
    else:

        if tkinter.messagebox.askyesno("エラー", "ファイルが選択されてないよ！今ここで選択する？"):
            # ファイルダイアログを表示してファイルパスを取得
            file_path = tkinter.filedialog.askopenfilename(
                title="Excelファイルを選択してください",
                filetypes=[("Excel files", "*.xlsx *.xlsm")]
            )
            read_excel_file(file_path);
        else:
            tkinter.messagebox.showinfo("終了", "ファイルが選択されませんでした");
            print("ファイルが選択されませんでした");

def edit_excel_file_mass(progress_callback=None):
    global wb;
    global sheet_names;
    global ws;
    global file_path;
    global mass_number;

    mass_number_row = 9;
    header_end_row = 39;

    

    ws = wb[sheet_names[0]];

    col = ws["A"];    # A列を取得
    for cell in col:
        if cell.value == "測定質量数              : ":
            mass_number_row = cell.row;    # 行番号を取得
            print(f"質量数の行番号：{mass_number_row}");    # 行番号を表示
        if cell.value == "測定回数":
            header_end_row = cell.row;
            print(f"ヘッダーの終了行番号：{header_end_row}");    # 行番号を表示 
            break;    # ループを抜ける


    

    mass_number = ws[mass_number_row];# ラベル行目を取得
    print(type(mass_number));    # 取得した行の型を表示
    #print(f"質量数：{mass_number}");


    mass_number_listed = list(mass_number);    # セルの値を取得

    #mass_number_edited = [i for i in mass_number if type(i) == int];    # int型だけ残す
    print(mass_number_listed);    # int型の質量数を表示
    print(type(mass_number_listed));
    print(f"{mass_number_listed[0]=}");
    print(f"{mass_number_listed[0].value=}");    # セルの値を表示

    mass_number_excerpted = [cell.value for cell in mass_number_listed if type(cell.value) == int];
    print(f"{mass_number_excerpted=}");    # int型の質量数を表示

    ws.delete_rows(1, header_end_row);    # 1行目から39行目まで削除
    ws.delete_cols(1,1);
    ws.delete_cols(2,4);

    for cell in ws["A"]:
        cell.value = cell.value[1:12];    # A列の値をスライスして上書き

    ws.insert_rows(1, 1);    # 1行目に1行追加

    ws["A1"].value = "Elapsed Time (s)";
    for i in range(len(mass_number_excerpted)):
        ws.cell(row=1, column=i+2).value = "m=" + str(mass_number_excerpted[i]);    # 1行目に質量数を追加
    ws.delete_cols(len(mass_number_excerpted)+2, ws.max_column);

def save_excel_file():
    global wb;
    global sheet_names;
    global ws;
    global file_path;
    
    print("Excelファイルを保存します。");
    if wb is not None:
        dname = os.path.dirname(file_path);
        fname = os.path.basename(file_path);
        outputFilePath = dname + "/output/edited_" + fname;
        print(f"出力ファイルパス：{outputFilePath}");
        os.makedirs(dname + "/output", exist_ok=True);    # 出力先のディレクトリを作成
        wb.save(outputFilePath) # Excelファイルの保存
        print(f"Excelファイルが保存されました：{outputFilePath}");


    else:
        print("Excelファイルが読み込まれていません。先にread_excel_file()を実行してください。");

def excel_to_csv():
    global wb;
    global sheet_names;
    global ws;
    global file_path;

    excel_file = os.path.dirname(file_path) + "/output/edited_" + os.path.basename(file_path)
    csv_file = os.path.dirname(file_path) + "/output/edited_" + os.path.basename(file_path).replace('.xlsx', '.csv').replace('.xlsm', '.csv')

    
    # Excelファイルを読み込む
    df = pandas.read_excel(excel_file)
    
    # CSVファイルに書き込む
    df.to_csv(csv_file, index=False)

    """
    tkinter.Tk().withdraw()
    tkinter.messagebox.showinfo('メッセージ', "読み込んだxlsxをCSVに変換しました！/n(「output」フォルダに保存されています)")
    """
    print("読み込んだxlsxをCSVに変換しました！ /n (「output」フォルダに保存されています)")

def copy_command_for_Igor():

    csv_file_path_with_collon = os.path.dirname(file_path) + "/output/edited_" + os.path.basename(file_path).replace('.xlsx', '.csv').replace('.xlsm', '.csv')
    csv_file_path_with_collon = csv_file_path_with_collon.replace(":", "")
    csv_file_path_with_collon = csv_file_path_with_collon.replace("/", ":")
    print(f"{csv_file_path_with_collon=}");
    
    # クリップボードにコピー
    #pyperclip.copy('LoadWave/J/D/W/A/E=1/K=0 "D:DQM:学習:openpyxl:インスト:pythonOpenpyxlのまとめ:SelfCreate:Igor提携:output:edited_S1_241017_221354.csv"');
    pyperclip.copy(f'LoadWave/J/D/W/A/E=1/K=0 "{csv_file_path_with_collon}"');

def notify_user(message):
    import tkinter as tk
    from tkinter import messagebox

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    messagebox.showinfo('メッセージ', message, parent=root)
    root.destroy()

def xlsx_to_csv_to_igor_integrated(path=None, progress_callback=None, add_label_on_gui_callback=None):
    try:    
        global wb;
        global sheet_names;
        global ws;
        global file_path;
        global mass_number;
        global csv_file_path_with_collon;
        global date;
        global mass_number_excerpted;

        #-----------------.xlsxファイルの読み込み-----------------
        file_path = path;    # グローバル変数にファイルパスを格納
        """
        # ファイルダイアログを表示してファイルパスを取得
        file_path = tkinter.filedialog.askopenfilename(
            title="Excelファイルを選択してください",
            filetypes=[("Excel files", "*.xlsx *.xlsm")]
        )
        """

        # ファイルが選択された場合のみ処理
        if file_path:
            # Excelファイルを読み込む
            wb = openpyxl.load_workbook(file_path);

            sheet_names = wb.sheetnames;    # シート名のリストを取得
            
            #notify_user(f"{str(file_path)}'\n を読み込みます")
            #print(f"選択されたファイル：{file_path}");    # 選択されたファイル名を表示
                
            if progress_callback:
                progress_callback(50)


            for i, sheet_name in enumerate(sheet_names):
                ws = wb[sheet_name];    # シートを取得
                print(f"{i+1}番目のシート名：{sheet_name}");    # シート名を表示
                print("先頭セルの値：", ws.cell(row=1, column=1).value)
                print(f"最大行数：{ws.max_row}");    # A列の行数を表示

            if progress_callback:
                progress_callback(55)

        else:

            if tkinter.messagebox.askyesno("エラー", "ファイルが選択されてないよ！今ここで選択する？"):
                # ファイルダイアログを表示してファイルパスを取得
                file_path = tkinter.filedialog.askopenfilename(
                    title="Excelファイルを選択してください",
                    filetypes=[("Excel files", "*.xlsx *.xlsm")]
                )
                read_excel_file(file_path);
            else:
                tkinter.messagebox.showinfo("終了", "ファイルが選択されませんでした");
                print("ファイルが選択されませんでした");

        #-----------------MASS整形-----------------
        mass_number_row = 9;
        header_end_row = 39;
        date_row = 31;

        

        ws = wb[sheet_names[0]];

        col = ws["A"];    # A列を取得
        for cell in col:
            if cell.value == "測定質量数              : ":
                mass_number_row = cell.row;    # 行番号を取得
                print(f"質量数の行番号：{mass_number_row}");    # 行番号を表示
            if cell.value == "測定開始日時            : ":
                date_row = cell.row;
                print(f"測定開始日時の行番号：{date_row}");    # 行番号を表示
            if cell.value == "測定回数":
                header_end_row = cell.row;
                print(f"ヘッダーの終了行番号：{header_end_row}");    # 行番号を表示 
                break;    # ループを抜ける
        if progress_callback:
            progress_callback(60)
        
        date_n_time = ws.cell(row=date_row, column=2).value;# 測定開始日時を取得
        date = (date_n_time.split()[0]).replace("/","");# 日付を取得
        print(f"測定開始日時：{date}");    # 測定開始日時を表示


        mass_number = ws[mass_number_row];# ラベル行目を取得
        #print(type(mass_number));    # 取得した行の型を表示
        #print(f"質量数：{mass_number}");
            
         


        mass_number_listed = list(mass_number);    # セルの値を取得

        #mass_number_edited = [i for i in mass_number if type(i) == int];    # int型だけ残す
        #print(mass_number_listed);    # int型の質量数を表示
        #print(type(mass_number_listed));
        #print(f"{mass_number_listed[0]=}");
        #print(f"{mass_number_listed[0].value=}");    # セルの値を表示

        mass_number_excerpted = [cell.value for cell in mass_number_listed if type(cell.value) == int];
        print(f"{mass_number_excerpted=}");    # int型の質量数を表示

        ws.delete_rows(1, header_end_row);    # 1行目から39行目まで削除
        if progress_callback:
            progress_callback(63)
        
        ws.delete_cols(1,1);
        if progress_callback:
            progress_callback(66)

        ws.delete_cols(2,4);
        if progress_callback:
            progress_callback(70)
        if add_label_on_gui_callback:
            add_label_on_gui_callback(f"✅ファイル整形")  
        
        #i = 0;
        for cell in ws["A"]:
            cell.value = cell.value[1:12];    # A列の値をスライスして上書き
            #i += 1;
            #if progress_callback:#入れると進捗が遅くなるので入れなくていいや！
            #    progress_callback(70 + i/len(ws["A"])*10)
        if progress_callback:
            progress_callback(80)

        ws.insert_rows(1, 1);    # 1行目に1行追加

        if mass_number_listed[20].value ==10:
            print("最終列がm=10");
            
            #test_row_listed = list(ws[40]);
            #test_row_excerpted = [cell.value for cell in test_row_listed];
            #print(f"{test_row_excerpted=}");    
            ws.delete_cols(idx=22, amount=ws.max_column);  #引数に注意！最初と最後じゃないよ！
            ws.delete_cols(idx=len(mass_number_excerpted)+1, amount=21-len(mass_number_excerpted)-1);  #引数に注意！最初と最後じゃないよ！
        else:
            ws.delete_cols(idx=len(mass_number_excerpted)+2, amount=ws.max_column);
        

        ws["A1"].value = "Elapsed Time (s)"+ "_" + str(date);

        for i in range(len(mass_number_excerpted)):
            ws.cell(row=1, column=i+2).value = "m=" + str(mass_number_excerpted[i]) + "_" + str(date);    # 1行目に質量数を追加
            #if progress_callback:#入れると進捗が遅くなるので入れなくていいや！
            #    progress_callback(80 + i/len(mass_number_excerpted)*10)


        if progress_callback:
            progress_callback(90)


        #-----------------.xlsxファイルとして保存-----------------
        print("Excelファイルを保存します。");
        if wb is not None:
            dname = os.path.dirname(file_path);
            fname = os.path.basename(file_path);
            outputFilePath = dname + "/edited_" + fname; #.xlsx作成時(CSVtoxlsx04で作成済み)
            #outputFilePath = dname + "/output/edited_" + fname;
            print(f"出力ファイルパス：{outputFilePath}");
            #os.makedirs(dname + "/output", exist_ok=True);    # 出力先のディレクトリを作成
            wb.save(outputFilePath) # Excelファイルの保存
            print(f"Excelファイルが保存されました：{outputFilePath}");
            if progress_callback:
                progress_callback(93)

        else:
            print("Excelファイルが読み込まれていません。先にread_excel_file()を実行してください。");


        #-----------------.xlsxを.csvへ変換-----------------
        excel_file = outputFilePath;
        file_path = excel_file;    # グローバル変数にファイルパスを格納
        #excel_file = os.path.dirname(file_path) + "/output/edited_" + os.path.basename(file_path)
        csv_file = os.path.dirname(outputFilePath) +"/"+  os.path.basename(outputFilePath).replace('.xlsx', '.csv').replace('.xlsm', '.csv')

        
        # Excelファイルを読み込む
        df = pandas.read_excel(excel_file)
        if progress_callback:
            progress_callback(96)
        
        
        # CSVファイルに書き込む
        df.to_csv(csv_file, index=False)
        if progress_callback:
            progress_callback(100)



        """
        tkinter.Tk().withdraw()
        tkinter.messagebox.showinfo('メッセージ', "読み込んだxlsxをCSVに変換しました！/n(「output」フォルダに保存されています)")
        """
        print("読み込んだxlsxをCSVに変換しました！ /n (「output」フォルダに保存されています)")

        print(f"出力ファイルパス：{csv_file}");
        if add_label_on_gui_callback:
            add_label_on_gui_callback(f"✅CSV変換")  

        #-----------------Igorコマンドをクリップボードへ-----------------
        csv_file_path_with_collon = csv_file
        csv_file_path_with_collon = csv_file_path_with_collon.replace(":", "")
        csv_file_path_with_collon = csv_file_path_with_collon.replace("/", ":")
        """
        csv_file_path_with_collon = os.path.dirname(file_path) + "/output/edited_" + os.path.basename(file_path).replace('.xlsx', '.csv').replace('.xlsm', '.csv')
        csv_file_path_with_collon = csv_file_path_with_collon.replace(":", "")
        csv_file_path_with_collon = csv_file_path_with_collon.replace("/", ":")
        """
        print(f"{csv_file_path_with_collon=}");
        
        # クリップボードにコピー
        #pyperclip.copy('LoadWave/J/D/W/A/E=1/K=0 "D:DQM:学習:openpyxl:インスト:pythonOpenpyxlのまとめ:SelfCreate:Igor提携:output:edited_S1_241017_221354.csv"');
        pyperclip.copy(f'LoadWave/J/D/W/A/E=1/K=0 "{csv_file_path_with_collon}"');
    except Exception as e:
        print("❌ excel_editor_01 エラー：", e)
        tb = traceback.extract_tb(e.__traceback__)
        last_trace = tb[-1]  # 最後のトレース（エラーが起きた場所）
        line_number = last_trace.lineno
        filename = last_trace.filename
        error_message = f"💥 不具合が起きたようです...以下の内容を小松路易に伝えてください:\n \n{e=}\n \n📄 ファイル: {filename}\n📍 行番号: {line_number}"
        messagebox.showerror("エラー発生！", error_message)
        return

def return_finalCSV_file_path():
    global file_path;
    os.path.basename(file_path)
    #tkinter.messagebox.showinfo('メッセージ', "return_xlsx_file_path");
    #print(file_path);
    return os.path.basename(file_path), os.path.dirname(file_path), file_path; # 返り値はファイル名とディレクトリ名,タプルで返す

def copy_to_clipboard():
    global csv_file_path_with_collon;
    global date;
    global mass_number_excerpted;

    #グラフへプロットするコマンド作成----------------------------------------------------
    Display_Wave_Command = " ";
    try:
        for i in range(len(mass_number_excerpted)):
            Display_Wave_Command += f"m_{mass_number_excerpted[i]}_{date}, ";      #f-stringでやる場合
        Display_Wave_Command += 'vs Elapsed_Time__s__' + str(date) + ' as "mass_' + str(date) + '"'; #普通にやる場合
        print(f"{Display_Wave_Command=}");
    except Exception as e:
        print("❌ excel_editor_01 エラー：", e)
        tb = traceback.extract_tb(e.__traceback__)
        last_trace = tb[-1]
        return


    #線の色・Table文字色を指定するコマンド作成-----------------------------------------------------
    Trace_Color_Command = "";
    try:
        excel_path = resource_path('個人用設定.xlsx');    # 個人用設定.xlsxのパスを取得
        wb = openpyxl.load_workbook(f'{excel_path}', data_only=True);    # 個人用設定.xlsxを読み込む 関数が合った場合は、data_only=Trueを指定して、計算結果のみを取得する
        ws = wb["TraceColor"];    # TraceColorを取得


        #↓だとlist(Trace_Color_Settings_mz)=[(1,), (2,), (3,)]になっちゃう
        #Trace_Color_Settings_mz = ws.iter_rows(min_row=2, max_row=max_row, max_col=1, values_only=True);
        Trace_Color_Settings_mz = ws["A"];    # TraceColorシートのA列を取得
        Trace_Color_Settings_mz_list = [cell.value for cell in Trace_Color_Settings_mz if type(cell.value) == int]; #数値だけ抽出してリストに格納

        Trace_Color_Settings_Color = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=4, values_only=True);   # TraceColorシートの色パラメータの範囲を取得
        Trace_Color_Settings_Color_tuple = tuple(Trace_Color_Settings_Color);    # タプルに変換
        #print(f'{Trace_Color_Settings_Color_tuple=}');

        #Trace_Color_Settings_dict
        #print(f'{Trace_Color_Settings_dict=}');

        #これだと上手くいかないなぁ
        #Trace_Color_Settings_dict = dict(zip(Trace_Color_Settings_mz_list, Trace_Color_Settings_Color));    # Trace_Color_Settingsを辞書に変換

        Trace_Color_Settings_dict = {mz:color for mz, color in zip(Trace_Color_Settings_mz_list, Trace_Color_Settings_Color_tuple) };    # 辞書に変換

        #print((Trace_Color_Settings_dict));

        #print(f'{Trace_Color_Settings_dict[1]=}');
        #print(f'{Trace_Color_Settings_dict[1][0]=}');
        for i in range(len(mass_number_excerpted)):
            Trace_Color_Command += f"•ModifyGraph rgb(m_{mass_number_excerpted[i]}_{date})={Trace_Color_Settings_dict[mass_number_excerpted[i]]};\n";
            Trace_Color_Command += f"•ModifyTable rgb(m_{mass_number_excerpted[i]}_{date})={Trace_Color_Settings_dict[mass_number_excerpted[i]]};\n";

    
    except Exception as e:
        print("❌ excel_editor_01 エラー：", e)
        tb = traceback.extract_tb(e.__traceback__)
        last_trace = tb[-1]
        return
    
    #Wave名・Wave表示名を指定するコマンド作成-----------------------------------------------------
    Wave_Rename_and_Retitle_Command = "";
    try:
        for i in range(len(mass_number_excerpted)):
            Wave_Rename_and_Retitle_Command += f'•ModifyTable title(m_{mass_number_excerpted[i]}_{date})="mz_{mass_number_excerpted[i]}_{date}";\n';
            Wave_Rename_and_Retitle_Command += f"•Rename m_{mass_number_excerpted[i]}_{date}, mz_{mass_number_excerpted[i]}_{date};\n";
        print(f"{Wave_Rename_and_Retitle_Command=}");
    except Exception as e:
        print("❌ excel_editor_01 エラー：", e)
        tb = traceback.extract_tb(e.__traceback__)
        last_trace = tb[-1]
        return

    #f-string内で改行するときは、\n使ってね！Igorでも反映されるよ！
    pyperclip.copy(f'//◆◆◆◆◆◆◆◆◆◆◆◆データ読み込み◆◆◆◆◆◆◆◆◆\nLoadWave/J/D/W/A/E=1/K=0 "{csv_file_path_with_collon}"\n•DoWindow/C/T mass_{date},"mass_{date}"\n//◆◆◆◆◆◆◆◆◆◆◆◆グラフへプロット◆◆◆◆◆◆◆◆◆◆◆◆\n//既存のグラフに追加したいなら Dispaly を AppendToGraph に書き換えてください.\n•Display {Display_Wave_Command}\n//◆◆◆◆◆◆◆◆◆◆◆◆LineColor変更◆◆◆◆◆◆◆◆◆◆\n{Trace_Color_Command}\n//◆◆◆◆◆◆◆◆◆◆◆◆LineSize変更◆◆◆◆◆◆◆◆◆◆◆\n•ModifyGraph lsize=1.5;\n//◆◆◆◆◆◆◆◆◆◆◆◆FontSize変更◆◆◆◆◆◆◆◆◆◆◆◆\n•ModifyGraph fSize=18;\n//◆◆◆◆◆◆◆◆◆◆◆◆Standoff,Mirror,FontSize◆◆◆◆◆◆◆\n•ModifyGraph tick=2,mirror=1,fSize=18,standoff=0;\n//◆◆◆◆◆◆◆◆◆◆◆◆Axis◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆\n•ModifyGraph log(left)=0;•ModifyGraph prescaleExp(left)=12;\n•ModifyGraph prescaleExp(bottom)=-3;\n•ModifyGraph prescaleExp(bottom)=-3;\n•ModifyGraph axisOnTop=1;\n•ModifyGraph dateInfo(bottom)={{1,0,2}};\nSetAxis/A=2/N=1 left;\n//◆◆◆◆◆◆◆◆◆◆◆◆AxisLabel◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆\n•Label left "\\Z24MASS signal intensity (pA)";\n•Label bottom "\\Z20Time (ks)";\n•ModifyGraph ZisZ=1;\n//◆◆◆◆◆◆◆◆◆◆◆◆Legend◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆\n•Legend/C/N=text0/F=0/B=1/M/LS=2;\n//--------------------Wave Rename&Retitle--------------------\n//任意の名称に変更することもできます．\n//•ModifyTable title(Wave名)="任意のWave表示名"	...WaveのTable上での表示名を変更します.例えばLegendやDatabrowserには影響しません.\n//•Rename 元のWave名,任意のWave名						...Waveの名前を根本から変更します.LegendやDatabrowserに影響します.アンダーバー以外の特殊な記号を名前に含める場合は,\'任意のWave名\'のようにシングルクオーテーションで囲んでください.\n{Wave_Rename_and_Retitle_Command}\n//手動でAxisのModeを "Date/Time" から "Linear" に変更してください.(これだけなぜかコマンドから制御できない...)');


def resource_path(filename: str):
    """PyInstaller対応：実行ファイルからのパスを解決する関数"""
    if getattr(sys, 'frozen', False):  # .exeとして実行中なら
        base_path = sys._MEIPASS       # PyInstallerの展開先フォルダ
    else:
        base_path = os.path.dirname(__file__)  # スクリプトとして実行中
    return os.path.join(base_path, filename)